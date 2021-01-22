from openpyxl import load_workbook

exec(open("convert_csv.py").read())


arkgWB = load_workbook(filename = 'arkg_holdings.xlsx') 
arkgSheet = arkgWB.active

sharesWB = load_workbook(filename = 'shareschange.xlsx')
sharesSheet = sharesWB.active

insertAtCol = 6
sharesSheet.insert_cols(idx=insertAtCol)
sharesSheet["F1"] = arkgSheet.cell(row=2, column=1).value

counter = 2 
while counter < arkgSheet.max_row:
    found = False
    stockName = arkgSheet.cell(row=counter, column=4).value
    for cell in sharesSheet['B']:
        if cell.value == stockName:
            sharesSheet.cell(row=cell.row, column=insertAtCol).value = arkgSheet.cell(row=counter, column=6).value
            found = True
    # add new stock listing to the bottom of the list
    if not found:
        sharesSheet.insert_rows(idx=2)
        sharesSheet.cell(row=2, column=1).value = stockName
        sharesSheet.cell(row=2, column=2).value = arkgSheet.cell(row=counter, column = 4).value
        sharesSheet.cell(row=2, column=insertAtCol).value = arkgSheet.cell(row=counter, column=6).value
        # sharesSheet.append({1 : stockName, 2 : arkgSheet.cell(row=counter, column = 4).value, 4:arkgSheet.cell(row=counter, column=6).value})
    counter+=1


sharesWB.save(filename="shareschange.xlsx")

